"""Score import helpers for Classwork CSV/XLSX results."""

import csv
import io
from dataclasses import dataclass
from pathlib import Path


ALLOWED_EXTENSIONS = {"csv", "xlsx"}
MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_ROWS = 2000


@dataclass
class ImportRow:
    raw: dict
    name: str | None
    email: str | None
    student_id: str | None
    score: float | None


def allowed_import(filename: str) -> bool:
    return "." in filename and Path(filename).suffix.lower().lstrip(".") in ALLOWED_EXTENSIONS


def _clean(value):
    if value is None:
        return ""
    return str(value).strip()


def _header_key(value):
    return " ".join(_clean(value).lower().replace("_", " ").split())


def detect_columns(headers):
    normalized = {_header_key(header): header for header in headers}
    aliases = {
        "name": ["name", "student name", "full name", "student"],
        "email": ["email", "student email", "school email", "email address"],
        "student_id": ["student id", "student number", "student no", "id"],
        "score": ["score", "points", "grade", "total score", "points earned"],
        "max_score": ["max score", "maximum score", "total points", "max points"],
    }
    found = {}
    for field, choices in aliases.items():
        for choice in choices:
            if choice in normalized:
                found[field] = normalized[choice]
                break
    return found


def parse_number(value):
    value = _clean(value).replace("%", "")
    if not value:
        return None
    try:
        return float(value)
    except ValueError:
        return None


def parse_csv(data: bytes):
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    rows = list(reader)
    if len(rows) > MAX_ROWS:
        raise ValueError(f"Import is limited to {MAX_ROWS} rows.")
    return headers, rows


def parse_xlsx(data: bytes):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("XLSX import requires openpyxl to be installed.") from exc

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    values = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not values:
        return [], []
    headers = ["" if value is None else str(value).strip() for value in values[0]]
    rows = []
    for values_row in values[1:MAX_ROWS + 2]:
        if not any(value not in (None, "") for value in values_row):
            continue
        rows.append({headers[index]: values_row[index] if index < len(values_row) else "" for index in range(len(headers))})
    if len(rows) > MAX_ROWS:
        raise ValueError(f"Import is limited to {MAX_ROWS} rows.")
    return headers, rows


def parse_file(filename: str, data: bytes):
    if len(data) > MAX_IMPORT_BYTES:
        raise ValueError("Import file must be 5 MB or smaller.")
    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        return parse_csv(data)
    if extension == ".xlsx":
        return parse_xlsx(data)
    raise ValueError("Only CSV and XLSX files are supported.")


def normalize_rows(headers, raw_rows):
    columns = detect_columns(headers)
    rows = []
    for raw in raw_rows:
        rows.append(
            ImportRow(
                raw=raw,
                name=_clean(raw.get(columns.get("name"))) if columns.get("name") else None,
                email=_clean(raw.get(columns.get("email"))).lower() if columns.get("email") else None,
                student_id=_clean(raw.get(columns.get("student_id"))) if columns.get("student_id") else None,
                score=parse_number(raw.get(columns.get("score"))) if columns.get("score") else None,
            )
        )
    return columns, rows
